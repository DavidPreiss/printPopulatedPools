#Copies values from one xlsx file to another
print("MASTERFILE START")
# Hard-coded values

SOURCE_FILE_PATH = "workdata/2023'.xlsx"
SOURCE_SHEET_NAME = "December 2023"

TARGET_FILE_PATH = "workdata/North.xlsx"
TARGET_SHEET_NAME = "N-2 "

SOURCE_START_ROW = 37
SOURCE_START_COLUMN = "G"

SOURCE_END_ROW = 72
SOURCE_END_COLUMN = "AM"

TARGET_START_ROW = 10
TARGET_START_COLUMN = "M"

NUMBER_OF_ROWS_IN_N_1 = 30
NUMBER_OF_ROWS_IN_N_2 = 36
NUMBER_OF_ROWS_IN_N_3 = 26

FINAL_OUTPUT_PATH = "MASTERFILE_OUTPUT.pdf"
TEMP_TARGET_FILE_PATH = "North_temp.xlsx"

import shutil
import os
import subprocess

try:
    import openpyxl
except ImportError as e:
    print(f"Error: {e}\nopenpyxl is not installed. Installing...")
    subprocess.check_call(["pip", "install", "openpyxl"])
    print("Installation complete. You can now run the script.")
    exit()
try:
    import PyPDF2
except ImportError as e:
    print(f"Error: {e}\nPyPDF2 is not installed. Installing...")
    subprocess.check_call(["pip", "install", "PyPDF2"])
    print("Installation complete. You can now run the script.")
    exit()

def column_letter_to_number(column_letter):
    """
    Convert Excel-style column letters to column numbers.
    Example: A -> 1, Z -> 26, AA -> 27, AB -> 28, ...
    """
    column_number = 0
    for char in column_letter:
        column_number = column_number * 26 + ord(char) - ord('A') + 1
    return column_number

def copy_excel_file(source_path, destination_path):
    try:
        shutil.copy2(source_path, destination_path)
        print(f'Successfully copied {source_path} to {destination_path}.')
    except Exception as e:
        print(f'An error occurred: {e}')

def copy_paste_cells(src_file_path, src_sheet_name, src_start_row, src_start_col, src_end_row, src_end_col,
                     target_file_path, target_sheet_name, target_start_row, target_start_col):
    try:
        # Load source workbook
        src_wb = openpyxl.load_workbook(src_file_path)
        
        # Check if source sheet exists, if not, use the active sheet
        if src_sheet_name not in src_wb.sheetnames:
            print(f"Warning: Source sheet '{src_sheet_name}' not found.")
            print(f"Using active sheet '{src_wb.active}'")
            src_ws = src_wb.active
        else:
            src_ws = src_wb[src_sheet_name]

        # Check if target path is the same as source path
        if target_file_path == src_file_path:
            # If target path is the same, assume target is in the same workbook
            target_wb = src_wb
            target_ws = src_wb[target_sheet_name]
        else:
            # If target path is different, load target workbook
            target_wb = openpyxl.load_workbook(target_file_path)
            
            # Check if target sheet exists, if not, use the active sheet
            if target_sheet_name not in target_wb.sheetnames:
                print(f"Warning: Target sheet '{target_sheet_name}' not found.")
                print(f"Using active sheet '{target_wb.active}'")
                target_ws = target_wb.active
            else:
                target_ws = target_wb[target_sheet_name]

        # Copy cells from source to target
        for row in range(src_start_row, src_end_row + 1):
            for col in range(src_start_col, src_end_col + 1):
                cell_value = src_ws.cell(row=row, column=col).value
                target_ws.cell(row=row - src_start_row + target_start_row, column=col - src_start_col + target_start_col).value = cell_value

        # Save the target workbook
        target_wb.save(target_file_path)

        print("Cells copied from source to target successfully!")

    except FileNotFoundError:
        print(f"Error: The file '{src_file_path}' or '{target_file_path}' was not found.")
    except PermissionError:
        print("Error: Permission issue. Make sure you have the necessary permissions to access the files.")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")

def find_empty_cells(file_path, columns_to_check, max_rows_to_check, start_row):
    
    try:
        print("opening "+file_path)
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {file_path}")

    empty_rows = []

    for column in columns_to_check:
        sheet = workbook.active
        print("checking column "+str(column)+" starting at row "+str(start_row))

        for row_num in range(start_row, max_rows_to_check + 1):
            cell_value = sheet.cell(row=row_num, column=column).value

            if cell_value is None or cell_value == "":
                empty_rows.append(row_num - start_row + 1)
                print(f"row '{row_num}' col '{column}' is empty")

    return empty_rows

def excel_to_pdf_with_libreoffice(excel_file_path, output_pdf_name, soffice_path):
    try:
        # Ensure that the output PDF directory exists (current working directory)
        output_dir = os.getcwd()
        os.makedirs(output_dir, exist_ok=True)
        print(f"output dir: '{output_dir}'")
        
        # Provide the full path to the soffice executable
        # soffice_path = r"C:/Program Files/LibreOffice/program/soffice.exe"

        # Use subprocess to run LibreOffice in headless mode for conversion
        subprocess.run([soffice_path, "--headless", "--convert-to", "pdf", "--outdir", output_dir, excel_file_path])
        
        
        # Create the full path for the output PDF
        file_name = os.path.basename(excel_file_path)
        file = os.path.splitext(file_name)
        output_pdf_path = os.path.join(output_dir ,(file[0] + ".pdf"))
        
        print(f"PDF created successfully: {output_pdf_path}")

        return output_pdf_path

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

def pdf_to_pdf_exclude_pages(input_path, output_path, list_excluded_pages):

    
    with open(input_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        pdf_writer = PyPDF2.PdfWriter()

        for page_number in range(len(pdf_reader.pages)):
            if page_number + 1 not in list_excluded_pages:
                page = pdf_reader.pages[page_number]
                pdf_writer.add_page(page)

        with open(output_path, 'wb') as output_file:
            pdf_writer.write(output_file)

# Convert column variables to integers if they are strings
if isinstance(SOURCE_START_COLUMN, str):
    SOURCE_START_COLUMN = column_letter_to_number(SOURCE_START_COLUMN)
if isinstance(SOURCE_END_COLUMN, str):
    SOURCE_END_COLUMN = column_letter_to_number(SOURCE_END_COLUMN)
if isinstance(TARGET_START_COLUMN, str):
    TARGET_START_COLUMN = column_letter_to_number(TARGET_START_COLUMN)

# Save Target File as a temp file for modification
copy_excel_file(TARGET_FILE_PATH, TEMP_TARGET_FILE_PATH)


# Display information about the copy-paste operation
print(f"Attempting to Copy cells ({SOURCE_START_ROW}, {SOURCE_START_COLUMN}) "
      f"to ({SOURCE_END_ROW}, {SOURCE_END_COLUMN}) in '{SOURCE_FILE_PATH}' sheet '{SOURCE_SHEET_NAME}'")
print(f"And Paste into cells ({TARGET_START_ROW}, {TARGET_START_COLUMN}) "
      f"to ({TARGET_START_ROW + (SOURCE_END_ROW - SOURCE_START_ROW)}, "
      f"{TARGET_START_COLUMN + (SOURCE_END_COLUMN - SOURCE_START_COLUMN)}) "
      f"in '{TEMP_TARGET_FILE_PATH}' sheet '{TARGET_SHEET_NAME}'")


# Call copy_paste_cells with the converted values and global row values
copy_paste_cells(SOURCE_FILE_PATH, SOURCE_SHEET_NAME, SOURCE_START_ROW, 
                 SOURCE_START_COLUMN, SOURCE_END_ROW, SOURCE_END_COLUMN, 
                 TEMP_TARGET_FILE_PATH, TARGET_SHEET_NAME, TARGET_START_ROW, TARGET_START_COLUMN)

# Find the rows that correspond to empty cells in the PH column
print(f"Checking for empty cells in '{TEMP_TARGET_FILE_PATH}'")
columns_to_check = [TARGET_START_COLUMN]
list_empty_rows = find_empty_cells(TEMP_TARGET_FILE_PATH, columns_to_check, (SOURCE_END_ROW - SOURCE_START_ROW), TARGET_START_ROW)
print(list_empty_rows)
print("END")

# Convert the .xlsx file to a pdf
print(f"Attempting to convert '{TEMP_TARGET_FILE_PATH}' into a pdf file")

# Provide the full path to the soffice executable
input_soffice_path = r"C:/Program Files/LibreOffice/program/soffice.exe"
raw_pdf_path = excel_to_pdf_with_libreoffice(TEMP_TARGET_FILE_PATH, "Masterfile.pdf", input_soffice_path)

print(f"Created file: '{raw_pdf_path}'")

# find the pdf and create a version stripped of the false pages
# 8 and 18 need to turn into pages 41 and 51
# for N-2, pages are the row numbers + NUMBER_OF_ROWS_IN_N_1 + 3
list_excluded_pages = []
#First add all the excluded pages for N-1

#Then add all excluded pages for N-2
for row in list_empty_rows:
    list_excluded_pages.append(row + NUMBER_OF_ROWS_IN_N_1 + 3)
print(f"The excluded pages list is: '{list_excluded_pages}'")

#Then add all excluded pages for N-3

#Then convert the pdf file into one that doesnt have the excluded pages
pdf_to_pdf_exclude_pages(raw_pdf_path, FINAL_OUTPUT_PATH, list_excluded_pages)

print(f"Created file: '{FINAL_OUTPUT_PATH}'")
# Prompt the user to press Enter before closing
input("Press Enter to close...")