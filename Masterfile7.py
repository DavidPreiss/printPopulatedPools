#Created by David Preiss
print("\nMASTERFILE7 START\n")
# Hard-coded values

SOURCE_FILE_PATH = "2023'.xlsx"
SOURCE_SHEET_NAME = "December 2023"
COL_OF_CODES = 2

TARGET_FILE_PATH = "North2.xlsx"
TARGET_SHEET_NAME = "N-2 "

SOURCE_START_ROW = 37
SOURCE_START_COLUMN = "G"

SOURCE_END_ROW = 72
SOURCE_END_COLUMN = "AM"

TARGET_START_ROW = 10
TARGET_START_COLUMN = "M"

NUMBER_OF_ROWS_IN_N_1 = 30
NUMBER_OF_ROWS_IN_N_2 = 36
NUMBER_OF_ROWS_IN_N_3 = 26

FINAL_OUTPUT_PATH = "MASTERFILE7_OUTPUT.pdf"
TEMP_TARGET_FILE_PATH = "North_temp.xlsx"
EXTRA_PAGES_PER_SHEET = 3

BNR_LOGO_IMAGE_PATH = "B&R_Logo.png"

SKIP_COPY = True
#set to true if you don't mind directly modifying TARGET_FILE_PATH

import shutil
import os
import subprocess

try:
    import openpyxl
except ImportError as e:
    print(f"!--ERROR:{e}\nopenpyxl is not installed. Installing...")
    subprocess.check_call(["pip", "install", "openpyxl"])
    print("Installation complete. You can now run the script.")
    exit()
try:
    import PyPDF2
except ImportError as e:
    print(f"!--ERROR:{e}\nPyPDF2 is not installed. Installing...")
    subprocess.check_call(["pip", "install", "PyPDF2"])
    print("Installation complete. You can now run the script.")
    exit()

try:
    from datetime import datetime
except ImportError as e:
    print(f"!--ERROR:{e}\ndatetime is not installed. Installing...")
    subprocess.check_call(["pip", "install", "datetime"])
    print("Installation complete. You can now run the script.")
    exit()

try:
    import fitz
except ImportError as e:
    print(f"!--ERROR:{e}\nPyMuPDF is not installed. Installing...")
    subprocess.check_call(["pip", "install", "PyMuPDF"])
    print("Installation complete. You can now run the script.")
    exit()

try:
    import win32com.client
except ImportError as e:
    print(f"!--ERROR:{e}\nwin32com is not installed. Installing...")
    subprocess.check_call(["pip", "install", "pywin32"])
    print("Installation complete. You can now run the script.")
    exit()

def column_letter_to_number(column_letter):
    """
    Convert Excel-style column letters to column numbers.
    Example: A -> 1, Z -> 26, AA -> 27, AB -> 28, ...
    """
    column_number = 0
    for char in column_letter:
        column_number = column_number * 26 + ord(char) - ord('A') + 1
    return column_number

def copy_xlsx_file(source_path, destination_path):
    try:
        
        print(f'Attempting to copy {source_path} to {destination_path}.')
        
        print(f"Current Working Directory: {os.getcwd()}")
        if not os.access(source_path, os.R_OK):
            print(f"!--ERROR: Cannot read {source_path}")
        if not os.access(destination_path, os.W_OK):
            print(f"!--ERROR: Cannot write to {destination_path}")


        shutil.copy2(source_path, destination_path)
        print(f'Successfully copied {source_path} to {destination_path}.')
    except Exception as e:
        print(f'!--ERROR occurred in copy_xlsx_file(): {e}')

def copy_paste_cells(src_file_path, src_sheet_name, src_start_row, src_start_col, src_end_row, src_end_col,
                     target_file_path, target_sheet_name, target_start_row, target_start_col):
    try:
        # Load source workbook
        src_wb = openpyxl.load_workbook(src_file_path)
        
        # Check if source sheet exists, if not, use the active sheet
        if src_sheet_name not in src_wb.sheetnames:
            print(f"Warning: Source sheet '{src_sheet_name}' not found.")
            print(f"Using active sheet '{src_wb.active}'")
            src_ws = src_wb.active
        else:
            src_ws = src_wb[src_sheet_name]

        # Check if target path is the same as source path
        if target_file_path == src_file_path:
            # If target path is the same, assume target is in the same workbook
            target_wb = src_wb
            target_ws = src_wb[target_sheet_name]
        else:
            # If target path is different, load target workbook
            target_wb = openpyxl.load_workbook(target_file_path)
            
            # Check if target sheet exists, if not, use the active sheet
            if target_sheet_name not in target_wb.sheetnames:
                print(f"Warning: Target sheet '{target_sheet_name}' not found.")
                print(f"Using active sheet '{target_wb.active}'")
                target_ws = target_wb.active
            else:
                target_ws = target_wb[target_sheet_name]

        # Copy cells from source to target
        for row in range(src_start_row, src_end_row + 1):
            for col in range(src_start_col, src_end_col + 1):
                cell_value = src_ws.cell(row=row, column=col).value
                target_ws.cell(row=row - src_start_row + target_start_row, column=col - src_start_col + target_start_col).value = cell_value

        # Save the target workbook
        target_wb.save(target_file_path)
        target_wb.close()
        src_wb.close()
        print("Cells copied from source to target successfully!")

    except FileNotFoundError:
        print(f"!--ERROR:The file '{src_file_path}' or '{target_file_path}' was not found.")
    except PermissionError:
        print("!--ERROR:Permission issue. Make sure you have the necessary permissions to access the files.")
    except Exception as e:
        print(f"!--ERROR occurred in copy_paste_cells(): {str(e)}")

def find_empty_cells(file_path, sheet_to_check, columns_to_check, max_rows_to_check, start_row):
    
    #Returns list of empty cells in given column
    #number in list represents order within cells checked
    # for example if you skip the first 10 rows and start checking on row 11,
    #   and row 12 is empty, then the first number in the list will be 2
    #   because its the second cell that was checked
    #   see comment within code to see where it happens vvvv
    try:
        print("opening "+file_path)
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {file_path}")
    except Exception as e:
        print(f"!--ERROR occurred in find_empty_cells(): {str(e)}")

    empty_rows = []
    

    for column in columns_to_check:
        print(f"checking column {str(column)} starting at row {str(start_row)} for a max of {max_rows_to_check} rows")
        print(f"Attempting to open workbook {file_path} on sheet {sheet_to_check}")
        sheet = workbook[sheet_to_check]
        print(f"Successfully opened workbook {file_path} on sheet {sheet_to_check}")

        for row_num in range(start_row, max_rows_to_check + start_row+1):
            cell_value = sheet.cell(row=row_num, column=column).value

            if cell_value is None or cell_value == "":
                empty_rows.append(row_num - start_row + 1) #this is where the conversion happens
                print(f"row '{row_num}' col '{column}' is empty")

    workbook.close()
    return empty_rows

def xlsx_to_pdf_with_libreoffice(xlsx_file_path, output_pdf_name):
    try:
        # Ensure that the output PDF directory exists (current working directory)
        output_dir = os.getcwd()
        os.makedirs(output_dir, exist_ok=True)
        print(f"output dir: '{output_dir}'")
        
        # Provide the full path to the soffice executable
        soffice_path = r"C:/Program Files/LibreOffice/program/soffice.exe"

        # Use subprocess to run LibreOffice in headless mode for conversion
        subprocess.run([soffice_path, "--headless", "--convert-to", "pdf", "--outdir", output_dir, xlsx_file_path])
        
        
        # Create the full path for the output PDF
        file_name = os.path.basename(xlsx_file_path)
        file = os.path.splitext(file_name)
        output_pdf_path = os.path.join(output_dir ,(file[0] + ".pdf"))
        
        print(f"PDF created with libre successfully: {output_pdf_path}")

        return output_pdf_path

    except Exception as e:
        print(f"!--ERROR occurred in xlsx_to_pdf_with_libreoffice(): {str(e)}")
        return None

def xlsx_to_pdf_with_excel(xlsx_file_path, output_pdf_name):
    try:
        # Ensure that the output PDF directory exists (current working directory)
        output_dir = os.getcwd()
        os.makedirs(output_dir, exist_ok=True)
        print(f"Output directory: '{output_dir}'")

        # Connect to Excel application
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False

        # Open the xlsx file
        workbook = excel_app.Workbooks.Open(os.path.abspath(xlsx_file_path))

        # Create the full path for the output PDF
        file_name = os.path.basename(xlsx_file_path)
        file = os.path.splitext(file_name)
        output_pdf_path = os.path.join(output_dir, (file[0] + ".pdf"))

        # Export the workbook to PDF
        workbook.ExportAsFixedFormat(0, output_pdf_path)

        # Close the workbook and quit Excel
        workbook.Close(False)
        excel_app.Quit()

        print(f"PDF created with excel successfully: {output_pdf_path}")
        return output_pdf_path

    except Exception as e:
        print(f"!--ERROR occurred in xlsx_to_pdf_with_excel(): {str(e)}")
        return None

def pdf_to_pdf_exclude_pages(input_path, output_path, list_excluded_pages):

    
    with open(input_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        pdf_writer = PyPDF2.PdfWriter()

        for page_number in range(len(pdf_reader.pages)):
            if page_number + 1 not in list_excluded_pages:
                page = pdf_reader.pages[page_number]
                pdf_writer.add_page(page)

        with open(output_path, 'wb') as output_file:
            pdf_writer.write(output_file)

def find_matching_cells(file_path, target_string, column_number):
    
    #Takes path of .xlsx file, target_string, and a column_number as an int
    #finds and returns the row number of the first instance of target_string in column column_number
    #Also returns a string list of every cell underneath that instance, terminating on an empty cell
    
    try:
        print("Opening " + file_path)
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {file_path}")

    matching_rows = []
    content_list = []

    sheet = workbook.active

    for row_num in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_num, column=column_number).value

        # Check if the cell is not empty or contains only spaces
        if cell_value is not None and cell_value.strip() != "":
            # Remove spaces from both cell content and target string for comparison
            if cell_value.replace(" ", "") == target_string.replace(" ", ""):
                matching_rows.append(row_num)
                print(f"Match found for '{target_string.replace(" ", "")}' at row {row_num}, column {column_number}")

                # Collect content from cells underneath until an empty cell or a cell with only spaces is encountered
                next_row = row_num + 1
                while True:
                    next_cell_value = sheet.cell(row=next_row, column=column_number).value
                    if next_cell_value is None or next_cell_value.strip() == "":
                        break
                    content_list.append(str(next_cell_value))
                    next_row += 1

                break  # Stop after finding the first match
    workbook.close()
    
    # print(f"HEY! matching_rows: {matching_rows}")
    return matching_rows[0] if matching_rows else None, content_list

def iterate_through_sheets(xlsx_file_path):

    #iterates through sheets copying to TEMP_TARGET_FILE_PATH
    # returns list of excluded pages
    
    print(f"\niterate_through_sheets({xlsx_file_path}) START")
    try:
        # Load the xlsx workbook
        workbook = openpyxl.load_workbook(xlsx_file_path)

        # Get a list of sheet names
        sheet_names = workbook.sheetnames
        
        cycle_num = 0
        list_excluded_pages = []
        list_page_names = []
        total_previous_pages= 0

        # Iterate through sheets
        for sheet_name in sheet_names:
            # Access the sheet by name
            sheet = workbook[sheet_name]

            # Process the sheet as needed
            print(f"\nsheet: '{sheet}'")
            print(f"sheet_name: '{sheet_name}'")
            
            # find row# of Sheet data
            result_row, result_content = find_matching_cells(SOURCE_FILE_PATH, sheet_name, COL_OF_CODES)
            
            SOURCE_START_ROW = result_row + 1
            SOURCE_START_COLUMN = COL_OF_CODES+5

            SOURCE_END_ROW = SOURCE_START_ROW+ len(result_content)
            SOURCE_END_COLUMN = SOURCE_START_COLUMN+32

            TARGET_START_ROW = 10
            TARGET_START_COLUMN = column_letter_to_number("M")
            
            # Display information about the copy-paste operation
            print(f"Attempting to Copy cells ({SOURCE_START_ROW}, {SOURCE_START_COLUMN}) "
                  f"to ({SOURCE_END_ROW}, {SOURCE_END_COLUMN}) in '{SOURCE_FILE_PATH}' sheet '{SOURCE_SHEET_NAME}'")
            print(f"And Paste into cells ({TARGET_START_ROW}, {TARGET_START_COLUMN}) "
                  f"to ({TARGET_START_ROW + (SOURCE_END_ROW - SOURCE_START_ROW)}, "
                  f"{TARGET_START_COLUMN + (SOURCE_END_COLUMN - SOURCE_START_COLUMN)}) "
                  f"in '{xlsx_file_path}' sheet '{sheet_name}'")

            # Call copy_paste_cells with the converted values and global row values
            copy_paste_cells(SOURCE_FILE_PATH, SOURCE_SHEET_NAME, SOURCE_START_ROW, 
                             SOURCE_START_COLUMN, SOURCE_END_ROW, SOURCE_END_COLUMN, 
                             xlsx_file_path, sheet_name, TARGET_START_ROW, TARGET_START_COLUMN)
            
            # Find the rows that correspond to empty cells in the PH column
            print(f"Checking for empty cells in '{sheet_name}'")
            list_columns_to_check = [TARGET_START_COLUMN] #change this number to change the cloumn checked
            list_empty_rows = find_empty_cells(TEMP_TARGET_FILE_PATH, sheet_name, list_columns_to_check, (SOURCE_END_ROW - SOURCE_START_ROW), TARGET_START_ROW)
            print(f"Empty rows:{list_empty_rows}")
            
            
            # print(f" result_content: {result_content}")
            print(f"Adding result_content to list_page_names")
            # Check if result_content is not None before extending list_page_names
            if result_content is not None:
                list_page_names.extend(result_content)
            else:
                print("Result content is None. Skipping extension of list_page_names.")
            # add names for extra pages to list_page_names
            for sum_page_num in range(1, EXTRA_PAGES_PER_SHEET+1):
                list_page_names.append("Summary page of "+sheet_name+" page "+str(sum_page_num))
            #print(f"list_empty_rows: {list_empty_rows}")
            for row in list_empty_rows:
                list_excluded_pages.append(row + total_previous_pages)
            #print(f"The excluded pages list is: '{list_excluded_pages}'")
            
            total_previous_pages = total_previous_pages + len(result_content) + EXTRA_PAGES_PER_SHEET
        
        
        # Close the xlsx workbook
        workbook.close()
        
        print(f"\n iterate_through_sheets({xlsx_file_path}) END\n")
        # ret_list = []
        # ret_list.append(list_excluded_pages)
        # ret_list.append(list_page_names)
        # return ret_list
        return list_excluded_pages, list_page_names

    except Exception as e:
        print(f"!--ERROR occurred in iterate_through_sheets(): {str(e)}")

def split_pdf_pages(input_pdf_path, output_paths):
    # Check if the input PDF file exists
    if not os.path.exists(input_pdf_path):
        raise FileNotFoundError(f"Input PDF file not found: {input_pdf_path}")

    # Open the input PDF file
    with open(input_pdf_path, 'rb') as input_file:
        # Create a PDF reader object
        pdf_reader = PyPDF2.PdfReader(input_file)

        # Check if the number of pages in the input PDF matches the number of output paths
        if len(pdf_reader.pages) < len(output_paths):
            raise ValueError("Input PDF has fewer pages than elements in the output paths list.")

        # Create a folder with the current date and time as its name
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_folder = "split_pdf_pages_output_"+current_datetime
        os.makedirs(output_folder)
            
        # Iterate through pages and corresponding output paths
        for page_num, output_path in zip(range(len(output_paths)), output_paths):
            # Create a new PDF writer object
            pdf_writer = PyPDF2.PdfWriter()

            # Add the current page to the new PDF writer
            pdf_writer.add_page(pdf_reader.pages[page_num])

            # Save the new PDF to the specified output path
            output_file_path = os.path.join(output_folder, f"{output_path}.pdf")
            with open(output_file_path, 'wb') as output_file:
                pdf_writer.write(output_file)
            
            #print(f"Created '{output_file_path}.pdf'")
        # If there are more pages in the input PDF, print a warning
        if len(pdf_reader.pages) > len(output_paths):
            print("Warning: Input PDF has more pages than elements in the output paths list. "
                  "Subsequent pages will be ignored.")
    print(f"split_pdf_pages() END\n")

def paste_image_into_pdf(input_pdf_path, input_image_path, x1, y1, Width, Height, output_pdf_path):
    doc = fitz.open(input_pdf_path)
    rect = fitz.Rect(x1, y1, Width, Height)       # put thumbnail in upper left corner
    img = open(input_image_path, "rb").read()  # an image file
    img_xref = 0                         # first execution embeds the image
    for page in doc:
        img_xref = page.insert_image(rect, stream=img,
                     xref=img_xref # 2nd time reuses existing image
              )
    doc.save(output_pdf_path)

# Convert column variables to integers if they are strings
if isinstance(SOURCE_START_COLUMN, str):
    SOURCE_START_COLUMN = column_letter_to_number(SOURCE_START_COLUMN)
if isinstance(SOURCE_END_COLUMN, str):
    SOURCE_END_COLUMN = column_letter_to_number(SOURCE_END_COLUMN)
if isinstance(TARGET_START_COLUMN, str):
    TARGET_START_COLUMN = column_letter_to_number(TARGET_START_COLUMN)


# Change working directory to the script's directory
script_dir = os.path.dirname(os.path.realpath(__file__))
os.chdir(script_dir)

#Check if we're skipping the copy step
if SKIP_COPY:
    TEMP_TARGET_FILE_PATH = TARGET_FILE_PATH
else:
    # Save Target File as a temp file for modification
    copy_xlsx_file(TARGET_FILE_PATH, TEMP_TARGET_FILE_PATH)


# iterate through the sheets of the file

result = iterate_through_sheets(TEMP_TARGET_FILE_PATH)
print("\n All Sheets Read")
if result is not None:
    print("output was valid")
    list_excluded_pages, list_page_names = result

    # Convert the .xlsx file to a pdf
    print(f"Attempting to convert '{TEMP_TARGET_FILE_PATH}' into a pdf file with excel")

    # Provide the full path to the soffice executable
    raw_pdf_path = xlsx_to_pdf_with_excel(TEMP_TARGET_FILE_PATH, "Masterfile.pdf")

    print(f"Created file: '{raw_pdf_path}'")
    
    #Add pictures to the raw pdf
    print(f"Adding images...")
    image_pdf_path = "image_"+os.path.basename(raw_pdf_path)
    print(f"image_pdf_path: {image_pdf_path}")
    paste_image_into_pdf(raw_pdf_path, BNR_LOGO_IMAGE_PATH, 40, 0, 145, 145, image_pdf_path)

    #Then convert the pdf file into one that doesnt have the excluded pages
    pdf_to_pdf_exclude_pages(image_pdf_path, FINAL_OUTPUT_PATH, list_excluded_pages)

    print(f"Created file: '{FINAL_OUTPUT_PATH}'")
    
    # Then split each page of that pdf into their own pdfs and label them
    split_pdf_pages(image_pdf_path, list_page_names)

    # print(f"list_page_names:\n{list_page_names}")
else:
    print("DAMN")
# Prompt the user to press Enter before closing
print("\nMASTERFILE7 END\n")
input("Press Enter to close...")