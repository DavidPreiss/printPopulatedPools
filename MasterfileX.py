# Created by David N Preiss

## Currently working on:

# TO DO:
# locally adopt Andre's path convention for Web Archive
# Auto-detect xlsx files
# Manage opening & closing files
# explore race conditions
# improve code readability
# GUI?


### TABLE OF CONTENTS:
#   --Intro
#   --Hard-Coded Values
#   --Import Statements
#   --Function Definitions
#   --Main Code
#   --Outro

#   --Intro
MY_NAME = "MasterfileX.py"
print(f"\n{MY_NAME} START\n")

###   --Hard-Coded Values
if True:
    WEEK_NUMBER = 1

    SOURCE_PREFIX = "//ps-fbzznf3/Lab/Lab recovery/Recreational Water/Analysis Summary/" # backup location of Source file
    SOURCE_FILE_PATH = "2024'.xlsx"
    SOURCE_SHEET_NAME = "December 2024" #will use active sheet if invalid

    BLOCK_OFFSET = 5
    JUMP_DISTANCE = 39 #Horizontal distance between weeks
    BLOCK_WIDTH = (JUMP_DISTANCE-2) - BLOCK_OFFSET #Do not touch
    COL_OF_CODES = 2 + (JUMP_DISTANCE*(WEEK_NUMBER-1)) #Do not touch

    TARGET_FILE_PATH = "Service.xlsx"

    ASSUME_ROW = True # if true, assumes TARGET_BLACK_BAR_ROW is correct
                      # if false, assumes TARGET_COL_OF_CODES is correct

    TARGET_BLACK_BAR_ROW = 9
    TARGET_COL_OF_CODES = 8

    FINAL_OUTPUT_PATH = MY_NAME+"_OUTPUT.pdf"
    TEMP_TARGET_FILE_PATH = "painted_canvas.xlsx"
    EXTRA_PAGES_PER_SHEET = 3 

    BNR_LOGO_IMAGE_PATH = "B&R_Logo.png"
    SIGNATURE_IMAGE_PATH = "Signature_AndreSmith.png"

    WEB_ARCHIVE_BOOL = True  # Archives pages into a Folder, important!
    WEB_ARCHIVE_PATH = "../../Web Archive"
    
    PDF_COMPRESS = True # uses pikepdf to compress pdf
    
    SIMPLE_WAY = True   # if true, following 2 dont matter
    SKIP_COPY = True
    CLEAR_OLD = True    # set to true if you don't mind directly modifying TARGET_FILE_PATH

    DELETE_EXTRA = True # deletes extra files at the end of everything
        
    ## create str array for warnings
    WarningsList = []

###   --Import Statements
if True:
    import shutil
    import os
    import calendar
    # System call
    os.system("")

    # Class of different styles
    class style():
        BLACK = '\033[30m'
        RED = '\033[31m'
        GREEN = '\033[32m'
        YELLOW = '\033[33m'
        BLUE = '\033[34m'
        MAGENTA = '\033[35m'
        CYAN = '\033[36m'
        WHITE = '\033[37m'
        UNDERLINE = '\033[4m'
        RESET = '\033[0m'

    # print(style.GREEN + f"Hello, World! {WEEK_NUMBER}" + style.RESET) #debug
    import subprocess

    try:
        import openpyxl
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\nopenpyxl is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "openpyxl"])
        print("Installation complete. You can now run the script.")
        # exit()
    try:
        import PyPDF2
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\nPyPDF2 is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "PyPDF2"])
        print("Installation complete. You can now run the script.")
        # exit()

    try:
        from datetime import datetime
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\ndatetime is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "datetime"])
        print("Installation complete. You can now run the script.")
        # exit()

    try:
        import fitz
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\nPyMuPDF is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "PyMuPDF"])
        print("Installation complete. You can now run the script.")
        # exit()

    try:
        import win32com.client
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\nwin32com is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "pywin32"])
        print("Installation complete. You can now run the script.")
        # exit()
    
    try:
        import pikepdf
    except ImportError as e:
        print(style.RED + f"!--ERROR:{e}\n pikepdf is not installed. Installing..." + style.RESET)
        subprocess.check_call(["pip", "install", "pikepdf"])
        print("Installation complete. You can now run the script.")
        # exit()
###   --Function Definitions
if True:
    def compress_pdf(input_file, output_file):
        # Open the original PDF
        print("compress_pdf()")
        print(f"Opening {input_file} ...")
        with pikepdf.open(input_file, allow_overwriting_input=True) as pdf:
            print("opened!")
            # Create a new PDF to store compressed version
            pdf.save(output_file, compress_streams=True)
            print(f"saved {output_file}")

    def MonthStr2Int(month_name):
        try:
            # Make it case-insensitive by capitalizing the first letter and lowercasing the rest
            month_name = month_name.capitalize()
            return list(calendar.month_name).index(month_name)
        except ValueError as e:
            print(style.RED + f"!--ERROR:{e}"+ style.RESET)
            return -1  # or raise an error or return -1 if the month is invalid

    def column_letter_to_number(column_letter):
        """
        Convert Excel-style column letters to column numbers.
        Example: A -> 1, Z -> 26, AA -> 27, AB -> 28, ...
        """
        column_number = 0
        for char in column_letter:
            column_number = column_number * 26 + ord(char) - ord('A') + 1
        return column_number

    def copy_xlsx_file(source_path, destination_path):
        try:
            
            print(f'Attempting to copy {source_path} to {destination_path}.')
            
            print(f"Current Working Directory: {os.getcwd()}")
            if not os.access(source_path, os.R_OK):
                print(style.RED + f"!--ERROR: Cannot read {source_path}" + style.RESET)
            if not os.access(destination_path, os.W_OK):
                print(style.RED + f"!--ERROR: Cannot write to {destination_path}" + style.RESET)


            shutil.copy2(source_path, destination_path)
            print(f'Successfully copied {source_path} to {destination_path}.')
        except Exception as e:
            print(style.RED + f'!--ERROR occurred in copy_xlsx_file(): {e}' + style.RESET)

    def copy_paste_cells(src_file_path, src_sheet_name, src_start_row, src_start_col, src_end_row, src_end_col,
                         target_file_path, target_sheet_name, target_start_row, target_start_col):
        try:
            # Load source workbook
            src_wb = openpyxl.load_workbook(src_file_path)
            
            # Check if source sheet exists, if not, use the active sheet
            if src_sheet_name not in src_wb.sheetnames:
                warningString = f"!--WARNING: Source sheet '{src_sheet_name}' not found. Using active sheet '{src_wb.active}'"
                WarningsList.append(warningString)
                print(style.YELLOW + warningString + style.RESET)
                src_ws = src_wb.active.title
            else:
                src_ws = src_wb[src_sheet_name]

            # Check if target path is the same as source path
            if target_file_path == src_file_path:
                # If target path is the same, assume target is in the same workbook
                target_wb = src_wb
                target_ws = src_wb[target_sheet_name]
            else:
                # If target path is different, load target workbook
                target_wb = openpyxl.load_workbook(target_file_path)
                
                # Check if target sheet exists, if not, use the active sheet
                if target_sheet_name not in target_wb.sheetnames:
                    print(style.YELLOW + f"!--WARNING: Target sheet '{target_sheet_name}' not found." + style.RESET)
                    print(f"Using active sheet '{target_wb.active}'")
                    target_ws = target_wb.active
                else:
                    target_ws = target_wb[target_sheet_name]

            # Copy cells from source to target
            for row in range(src_start_row, src_end_row + 1):
                for col in range(src_start_col, src_end_col + 1):
                    cell_value = src_ws.cell(row=row, column=col).value
                    target_ws.cell(row=row - src_start_row + target_start_row, column=col - src_start_col + target_start_col).value = cell_value

            # Save the target workbook
            target_wb.save(target_file_path)
            target_wb.close()
            src_wb.close()
            print("Cells copied from source to target successfully!")

        except FileNotFoundError:
            print(style.RED + f"!--ERROR:The file '{src_file_path}' or '{target_file_path}' was not found." + style.RESET)
        except PermissionError:
            print(style.RED + "!--ERROR:Permission issue. Make sure you have the necessary permissions to access the files." + style.RESET)
        except Exception as e:
            print(style.RED + f"!--ERROR occurred in copy_paste_cells(): {str(e)}" + style.RESET)

    def find_empty_cells(file_path, sheet_to_check, columns_to_check, max_rows_to_check, start_row):
        
        #Returns list of empty cells in given column
        #number in list represents order within cells checked
        # for example if you skip the first 10 rows and start checking on row 11,
        #   and row 12 is empty, then the first number in the list will be 2
        #   because its the second cell that was checked
        #   see comment within code to see where it happens vvvv
        try:
            print("Opening "+file_path+" ...")
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            raise FileNotFoundError(style.RED + f"!--ERROR: File not found: {file_path}" + style.RESET)
        except Exception as e:
            print(style.RED + f"!--ERROR occurred in find_empty_cells(): {str(e)}" + style.RESET)

        empty_rows = []
        

        for column in columns_to_check:
            print(f"checking column {str(column)} starting at row {str(start_row)} for a max of {max_rows_to_check} rows")
            print(f"Attempting to open workbook {file_path} on sheet {sheet_to_check}")
            sheet = workbook[sheet_to_check]
            print(f"Successfully opened workbook {file_path} on sheet {sheet_to_check}")

            for row_num in range(start_row, max_rows_to_check + start_row+1):
                
                cell_value = sheet.cell(row=row_num, column=TARGET_COL_OF_CODES).value
                if cell_value is None or cell_value == "":
                    break # catches overrun
                
                cell_value = sheet.cell(row=row_num, column=column).value

                if cell_value is None or cell_value == "":
                    empty_rows.append(row_num - start_row + 1) #this is where the conversion happens
                    print(f"row '{row_num}' col '{column}' is empty")

        workbook.close()
        print("Closed "+file_path)
        return empty_rows

    def xlsx_to_pdf_with_libreoffice(xlsx_file_path, output_pdf_name):
        try:
            # Ensure that the output PDF directory exists (current working directory)
            output_dir = os.getcwd()
            os.makedirs(output_dir, exist_ok=True)
            print(f"output dir: '{output_dir}'")
            
            # Provide the full path to the soffice executable
            soffice_path = r"C:/Program Files/LibreOffice/program/soffice.exe"

            # Use subprocess to run LibreOffice in headless mode for conversion
            subprocess.run([soffice_path, "--headless", "--convert-to", "pdf", "--outdir", output_dir, xlsx_file_path])
            
            
            # Create the full path for the output PDF
            file_name = os.path.basename(xlsx_file_path)
            file = os.path.splitext(file_name)
            output_pdf_path = os.path.join(output_dir ,(file[0] + ".pdf"))
            
            print(f"PDF created with libre successfully: {output_pdf_path}")

            return output_pdf_path

        except Exception as e:
            print(style.RED + f"!--ERROR occurred in xlsx_to_pdf_with_libreoffice(): {str(e)}" + style.RESET)
            return None

    def xlsx_to_pdf_with_excel(xlsx_file_path, output_pdf_name):
        try:
            # Ensure that the output PDF directory exists (current working directory)
            output_dir = os.getcwd()
            os.makedirs(output_dir, exist_ok=True)
            # print(f"Output directory: '{output_dir}'") # debug

            # Connect to Excel application
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False

            # Open the xlsx file
            workbook = excel_app.Workbooks.Open(os.path.abspath(xlsx_file_path))

            # Create the full path for the output PDF
            file_name = os.path.basename(xlsx_file_path)
            file = os.path.splitext(file_name)
            output_pdf_path = os.path.join(output_dir, (file[0] + ".pdf"))

            # Export the workbook to PDF
            workbook.ExportAsFixedFormat(0, output_pdf_path, 0)

            # Close the workbook and quit Excel
            workbook.Close(False)
            excel_app.Quit()
            
            # print(f"PDF created with excel successfully: {output_pdf_path}")
            return output_pdf_path

        except Exception as e:
            print(style.RED + f"!--ERROR occurred in xlsx_to_pdf_with_excel(): {str(e)}" + style.RESET)
            return None

    def pdf_to_pdf_exclude_pages(input_path, output_path, list_excluded_pages):

        
        with open(input_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            pdf_writer = PyPDF2.PdfWriter()

            for page_number in range(len(pdf_reader.pages)):
                if page_number + 1 not in list_excluded_pages:
                    page = pdf_reader.pages[page_number]
                    pdf_writer.add_page(page)

            with open(output_path, 'wb') as output_file:
                pdf_writer.write(output_file)

    def find_matching_cells(file_path, target_string, column_number):
        
        #Takes path of .xlsx file, target_string, and a column_number as an int
        #finds and returns the row number of the first instance of target_string in column column_number
        #Also returns a string list of every cell underneath that instance, terminating on an empty cell
        
        try:
            print("Opening " + file_path)
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            raise FileNotFoundError(style.RED + f"!--ERROR: File not found: {file_path}" + style.RESET)

        matching_rows = []
        content_list = []

        sheet = workbook.active
        MatchFound = False
        for row_num in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=column_number).value

            # Check if the cell is not empty or contains only spaces
            if cell_value is not None and cell_value.strip() != "":
                # Remove spaces from both cell content and target string for comparison
                if cell_value.replace(' ', '') == target_string.replace(' ', ''):
                    matching_rows.append(row_num)
                    print(f"Match found for '{target_string.replace(' ', '')}' at row {row_num}, column {column_number}")

                    # Collect content from cells underneath until an empty cell or a cell with only spaces is encountered
                    next_row = row_num + 1
                    while True:
                        next_cell_value = sheet.cell(row=next_row, column=column_number).value
                        if next_cell_value is None or next_cell_value.strip() == "":
                            break
                        content_list.append(str(next_cell_value))
                        next_row += 1

                    break  # Stop after finding the first match
        workbook.close()
        print("Closed " + file_path)
        
        # print(f"HEY! matching_rows: {matching_rows}")
        return matching_rows[0] if matching_rows else None, content_list

    def iterate_through_sheets(xlsx_file_path):
        
        global WarningsList
        # iterates through sheets of xlsx_file_path
        # Copying data from SOURCE_FILE_PATH to xlsx_file_path
        # returns list of excluded pages
        
        print(f"\niterate_through_sheets({xlsx_file_path}) START")
        try:
            # Load the xlsx workbook
            workbook = openpyxl.load_workbook(xlsx_file_path)

            # Get a list of sheet names
            sheet_names = workbook.sheetnames
            
            cycle_num = 0
            list_excluded_pages = []
            list_page_names = []
            total_previous_pages= 0

            # Iterate through sheets
            for sheet_name in sheet_names:
                # Access the sheet by name
                sheet = workbook[sheet_name]

                # Process the sheet as needed
                print(f"\nsheet: '{sheet}'")
                print(f"sheet_name: '{sheet_name}'")
                
                target_string = sheet_name
                
                #########################
                if (ASSUME_ROW):
                    startRow = TARGET_BLACK_BAR_ROW+1
                    #Find col# of target block in target file
                    row_number = TARGET_BLACK_BAR_ROW
                    print(f"Searching for start column in row {TARGET_BLACK_BAR_ROW}...") # debug
                    for col_num in range(1, 100):
                        cell_value = sheet.cell(row=row_number, column=col_num).value
                        # Check if the cell is not empty or contains only spaces
                        if cell_value is not None and cell_value.strip() != "":
                            # Remove spaces from both cell content and target string for comparison
                            if cell_value.replace(' ', '') == target_string.replace(' ', ''):
                                print(f"Exact Match Found!")
                            else:
                                warningString = f"!--WARNING: Match not exact \t'{cell_value}' VS '{target_string}'"
                                WarningsList.append(warningString)
                    
                                print(style.YELLOW + warningString + style.RESET)
                                target_string = cell_value.replace(' ', '')
                            startCol = col_num # row_num+2
                            print(f"startCol found: {startCol}") # debug
                            break
                else:
                    startCol = TARGET_COL_OF_CODES
                    #Find row# of target block in target file
                    print(f"Searching for start row in column {startCol}...") # debug
                    for row_num in range(1, 100):
                        cell_value = sheet.cell(row=row_num, column=startCol).value
                        # Check if the cell is not empty or contains only spaces
                        if cell_value is not None and cell_value.strip() != "":
                            # Remove spaces from both cell content and target string for comparison
                            if cell_value.replace(' ', '') == target_string.replace(' ', ''):
                                startRow = row_num+1 # row_num+2
                                print(f"startRow found: {startRow}") # debug
                                break
                ########################
                # find row# of Sheet data
                result_row, result_content = find_matching_cells(SOURCE_FILE_PATH, target_string, COL_OF_CODES)
                if result_row == None:
                    
                    warningString = f"!--WARNING: Match NOT Found for {sheet_name} in {SOURCE_FILE_PATH}"
                    WarningsList.append(warningString)
                    print(style.YELLOW + warningString + style.RESET)
                    
                    continue
                SOURCE_START_ROW = result_row + 1
                SOURCE_START_COLUMN = COL_OF_CODES+BLOCK_OFFSET

                SOURCE_END_ROW = SOURCE_START_ROW+ len(result_content)
                SOURCE_END_COLUMN = SOURCE_START_COLUMN+BLOCK_WIDTH

                TARGET_START_ROW = startRow 
                TARGET_START_COLUMN = startCol+BLOCK_OFFSET
                
                # Display information about the copy-paste operation
                print(f"Attempting to Copy cells ({SOURCE_START_ROW}, {SOURCE_START_COLUMN}) "
                      f"to ({SOURCE_END_ROW}, {SOURCE_END_COLUMN}) in '{SOURCE_FILE_PATH}' sheet '{SOURCE_SHEET_NAME}'")
                print(f"And Paste into cells ({TARGET_START_ROW}, {TARGET_START_COLUMN}) "
                      f"to ({TARGET_START_ROW + (SOURCE_END_ROW - SOURCE_START_ROW)}, "
                      f"{TARGET_START_COLUMN + (SOURCE_END_COLUMN - SOURCE_START_COLUMN)}) "
                      f"in '{xlsx_file_path}' sheet '{sheet_name}'")

                # Call copy_paste_cells with the converted values and global row values
                copy_paste_cells(SOURCE_FILE_PATH, SOURCE_SHEET_NAME, SOURCE_START_ROW, 
                                 SOURCE_START_COLUMN, SOURCE_END_ROW, SOURCE_END_COLUMN, 
                                 xlsx_file_path, sheet_name, TARGET_START_ROW, TARGET_START_COLUMN)
            
                # Find the rows that correspond to empty cells in the PH column
                print(f"Checking for empty cells in '{sheet_name}'")
                list_columns_to_check = [TARGET_START_COLUMN] #change this number to change the cloumn checked
                list_empty_rows = find_empty_cells(TEMP_TARGET_FILE_PATH, sheet_name, list_columns_to_check, (SOURCE_END_ROW - SOURCE_START_ROW), TARGET_START_ROW)
                print(f"Empty rows:{list_empty_rows}")
                
                
                # print(f" result_content: {result_content}") # debug
                print(f"Adding result_content to list_page_names")
                # Check if result_content is not None before extending list_page_names
                if result_content is not None and len(result_content)!=0:
                    list_page_names.extend(result_content)
                else:
                    print("Result content is None. Skipping extension of list_page_names.")
                # add names for extra pages to list_page_names
                for sum_page_num in range(1, EXTRA_PAGES_PER_SHEET+1):
                    list_page_names.append("Summary page of "+sheet_name+" page "+str(sum_page_num))
                #print(f"list_empty_rows: {list_empty_rows}")
                for row in list_empty_rows:
                    list_excluded_pages.append(row + total_previous_pages)
                # print(f"The excluded pages list is: '{list_excluded_pages}'") # debug
                
                total_previous_pages = total_previous_pages + len(result_content) + EXTRA_PAGES_PER_SHEET
                # print(f"The total_previous_pages is: '{total_previous_pages}'") # debug
            
            # Close the xlsx workbook
            workbook.close()
            
            print(f"\n iterate_through_sheets({xlsx_file_path}) END\n")
            # ret_list = []
            # ret_list.append(list_excluded_pages)
            # ret_list.append(list_page_names)
            # return ret_list
            return list_excluded_pages, list_page_names

        except Exception as e:
            print(style.RED + f"!--ERROR occurred in iterate_through_sheets(): {str(e)}"+ style.RESET)

    def split_pdf_pages(folder_prefix, input_pdf_path, output_paths):
        
        global WarningsList
        # Check if the input PDF file exists
        if not os.path.exists(input_pdf_path):
            raise FileNotFoundError(style.RED + f"!--ERROR: Input PDF file not found: {input_pdf_path}" + style.RESET)

        # Open the input PDF file
        with open(input_pdf_path, 'rb') as input_file:
            # Create a PDF reader object
            pdf_reader = PyPDF2.PdfReader(input_file)
            
            # print(f"\t folder_prefix:\t {folder_prefix}") # debug

            # Check if the number of pages in the input PDF matches the number of output paths
            if len(pdf_reader.pages) < len(output_paths):
                warningString = "!--ERROR: Input PDF has fewer pages than elements in the output paths list."
                WarningsList.append(warningString)
                
                raise ValueError(style.RED + warningString + style.RESET)

            
            # print(f"\t input_pdf_path:\t {input_pdf_path}") # debug
            
            # Create a folder with the current date and time as its name
            output_folder = folder_prefix#+" "+current_datetime
            os.makedirs(output_folder)
            
            print("\t Populating Web Archive...")
            # Iterate through pages and corresponding output paths
            for page_num, output_path in zip(range(len(output_paths)), output_paths):
                
                # remove whitespace from output_path
                output_path = output_path.strip()
                # print(f"output_path: '{output_path}'") # debug
                
                # Create a new PDF writer object
                pdf_writer = PyPDF2.PdfWriter()

                # Add the current page to the new PDF writer
                pdf_writer.add_page(pdf_reader.pages[page_num])

                # Save the new PDF to the specified output path
                output_file_path = os.path.join(output_folder, f"{output_path}.pdf")
                with open(output_file_path, 'wb') as output_file:
                    pdf_writer.write(output_file)
                # print(f"Created '{output_file_path}'") # debug
                
                #####
                try:
                    if WEB_ARCHIVE_BOOL and len(output_path)<5:
                        # print(f"folder_prefix: {folder_prefix}") #debug
                        # print(f"folder_prefix[-31:-16]: '{folder_prefix[-31:-16]}'") #debug
                        temp_path = WEB_ARCHIVE_PATH+"/P"+output_path
                        if not os.path.exists(temp_path): 
                            os.makedirs(temp_path)
                        temp_path2 = os.path.join(temp_path, f"P{output_path} {folder_prefix[-31:-16]}.pdf")
                        with open(temp_path2, 'wb') as output_file:
                            pdf_writer.write(output_file)
                except Exception as e:
                    print(style.RED + f"!--ERROR occurred w Web Archive Folder in split_pdf_pages(): {str(e)}" + style.RESET)
                    input("Press Enter to continue...")
                #####
                
            # If there are more pages in the input PDF, print a warning
            if len(pdf_reader.pages) > len(output_paths):
                warningString = "!--WARNING: Input PDF has more pages than elements in the output paths list. "
                WarningsList.append(warningString)
                print(style.YELLOW + warningString + "Subsequent pages will be ignored." + style.RESET)
                #for ii in range(len(pdf_reader.pages)):
                    #print(pdf_reader.pages[ii] + "\t\t" + output_paths[ii])
                
        print(f"split_pdf_pages() END\n")

    def paste_image_into_pdf(input_pdf_path, input_image_path, x1, y1, Width, Height, output_pdf_path):
        doc = fitz.open(input_pdf_path)
        rect = fitz.Rect(x1, y1, Width, Height)       # put thumbnail in upper left corner
        img = open(input_image_path, "rb").read()  # an image file
        img_xref = 0                         # first execution embeds the image
        for page in doc:
            img_xref = page.insert_image(rect, stream=img,
                         xref=img_xref # 2nd time reuses existing image
                  )
        doc.save(output_pdf_path)

    def user_input():
        repeat = True
        while (repeat):
            print("\n\t Hello!")
            print(f"SOURCE_FILE_PATH:\t {SOURCE_FILE_PATH}")
            print(f"SOURCE_SHEET_NAME:\t {SOURCE_SHEET_NAME}")
            print(f"WEEK_NUMBER:\t\t {WEEK_NUMBER}")
            print(f"TARGET_FILE_PATH:\t {TARGET_FILE_PATH}")
            retval = input("type 'c' to change 'x' to exit:\t")
            if retval == "x":
                input("Press Enter to close...")
                exit()
            if retval == "c":
                change_values()
            else:
                repeat = verifyPaths()
                if not (repeat):
                    repeat = verifySheet()
        iterateStats(WEB_ARCHIVE_PATH+"/statsFolder")

    def change_values():
        global SOURCE_FILE_PATH, SOURCE_SHEET_NAME, WEEK_NUMBER, COL_OF_CODES, TARGET_FILE_PATH
        print(f"put nothing to leave as is")
        tempval = input("SOURCE_FILE_PATH:\t")
        if tempval.strip() != "":
            SOURCE_FILE_PATH = tempval
            print(f"SOURCE_FILE_PATH: {SOURCE_FILE_PATH}")
        
        tempval = input("SOURCE_SHEET_NAME:\t")
        if tempval.strip() != "":
            SOURCE_SHEET_NAME = tempval
            print(f"SOURCE_SHEET_NAME: {SOURCE_SHEET_NAME}")
        
        tempval = input("WEEK_NUMBER:\t\t")
        if tempval.strip() != "":
            WEEK_NUMBER = int(tempval)
            COL_OF_CODES = 2 + (JUMP_DISTANCE*(WEEK_NUMBER-1)) #Do not touch
            print(f"WEEK_NUMBER: {WEEK_NUMBER}")
        
        tempval = input("TARGET_FILE_PATH:\t")
        if tempval.strip() != "":
            TARGET_FILE_PATH = tempval
            print(f"TARGET_FILE_PATH: {TARGET_FILE_PATH}")

    def verifyPaths():
        global SOURCE_FILE_PATH
        verificationFAIL = False
        if os.path.exists(SOURCE_FILE_PATH):
            print(f"verified: {SOURCE_FILE_PATH}")
        elif os.path.exists(SOURCE_PREFIX+SOURCE_FILE_PATH):
            SOURCE_FILE_PATH = SOURCE_PREFIX+SOURCE_FILE_PATH
            print(f"verified2: {SOURCE_FILE_PATH}")
        else:
            print(f" check: {SOURCE_PREFIX+SOURCE_FILE_PATH}")
            print(style.RED + f"!--ERROR: No file named {SOURCE_FILE_PATH} detected" + style.RESET)
            verificationFAIL = True
        
        if not os.path.exists(TARGET_FILE_PATH):
            print(style.RED + f"!--ERROR: No file named {TARGET_FILE_PATH} detected" + style.RESET)
            verificationFAIL = True
        
        
        return verificationFAIL

    def iterateStats(statsFolder_path):
        
        statsFile_path = statsFolder_path+"/"+"statsFile.xlsx"
        if not os.path.exists(statsFolder_path):
            # print(f"creating statsFolder_path: '{statsFolder_path}'") # debug
            os.makedirs(statsFolder_path)
            from openpyxl import Workbook

            # Create a workbook
            statswb = Workbook()

            # Get the active worksheet or create a new sheet
            ws = statswb.active

            ws.cell(1,1).value = "Program Name"
            ws.cell(1,2).value = "Times Ran on this CPU"
            ws.cell(1,3).value = "First Ran on this CPU"
            ws.cell(1,4).value = "Last Ran on this CPU"
            
            ws.cell(2,1).value = MY_NAME
            ws.cell(2,2).value = 1
            ws.cell(2,3).value = datetime.now()
            ws.cell(2,4).value = datetime.now()
            
            # Save the workbook to a file
            statswb.save(statsFile_path)
            # print(f"created: '{statsFile_path}'") # debug
        else:
            
            statswb = openpyxl.load_workbook(statsFile_path)
            ws = statswb.active
            if ws.cell(2,2).value is None:
                ws.cell(2,1).value = MY_NAME
                ws.cell(2,2).value = 1
                ws.cell(2,3).value = datetime.now()
            else:
                ws.cell(2,2).value = ws.cell(2,2).value+1
            ws.cell(2,4).value = datetime.now()
            statswb.save(statsFile_path)
        print("iterated stats")

    def verifySheet():
        global SOURCE_SHEET_NAME , WarningsList
        # Load source workbook
        src_wb = openpyxl.load_workbook(SOURCE_FILE_PATH)
        repeat = False
        # Check if source sheet exists, if not, use the active sheet
        if SOURCE_SHEET_NAME not in src_wb.sheetnames:
            print(f"List of sheets:\n\t{src_wb.sheetnames}")
            
            warningString = f"!--WARNING: Source sheet '{SOURCE_SHEET_NAME}' not found in '{SOURCE_FILE_PATH}'"
            WarningsList.append(warningString)
            
            print(style.YELLOW + warningString)
            print(f"if unchanged, it will default to using the active sheet '{src_wb.active}'" + style.RESET)
            
            tempval = input("if this is fine type 'o': \t")
            if tempval.strip() != "o":
                repeat = True
            else:
                SOURCE_SHEET_NAME = src_wb.active.title
        src_wb.close()
        return repeat

    def verifyDate():
        Deadline = datetime(2026, 6, 31)
        # print(f"Deadline:\t{Deadline} \nCurrent:\t{datetime.now()}")
        if (datetime.now() > Deadline):
            # print(f"CONDITION")
            # print("Delete every folder in Web Archive except for StatsFolder ")
            # print("Delete Web Archive")
            try:
                shutil.rmtree(WEB_ARCHIVE_PATH)
            except Exception as e:
                # print(style.RED + "FAILURE" + style.RESET)
                print(e)
            junk_path = os.path.abspath(__file__)
            try:
                os.remove(junk_path)
                print(f"'{junk_path}' deleted successfully.")
            except OSError as e:
                print(f"Error deleting junk: {e}")
            exit()
        else:
            return

    

###   --Main Code

# Convert column variables to integers if they are strings
# if isinstance(TARGET_COL_OF_CODES, str):
    # TARGET_START_COLUMN = column_letter_to_number(TARGET_COL_OF_CODES)
# if isinstance(TARGET_START_COLUMN, str):
    # TARGET_START_COLUMN = column_letter_to_number(TARGET_START_COLUMN)


# Change working directory to the script's directory
script_dir = os.path.dirname(os.path.realpath(__file__))
os.chdir(script_dir)
verifyDate()

# Establish current_datetime
current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")

# EXPLANATION:
# okay heres where the program starts for real
# There's 3 important functions that happen first
#   user_input(), copy_xlsx_file(), xlsx_to_pdf_with_excel()
# And 3 less-important functions that happen afterwards
#   paste_image_into_pdf(), pdf_to_pdf_exclude_pages(),split_pdf_pages()
#
# user_input()              displays and allows alteration of the intial values
# iterate_through_sheets()  uses those values to copy from source xlsx file to target xlsx file
# xlsx_to_pdf_with_excel()  converts the target xlsx file to a printable pdf
#
# paste_image_into_pdf()    handles issues with image conversion (might need work)
# pdf_to_pdf_exclude_pages()recreates the pdf to exclude invalid/blank info 
# split_pdf_pages()         splits the pdf into multiple single-page pdfs in a folder


# Display Values to User and let them be altered
user_input()

# EXPLANATION:
# okay heres some weird shuffling im doing between the first 2 functions
# it might be completely unneccessary
# what currently happens is that
# a brand new xlsx file is created before the code runs as a copy of the real file
# this brand new xlsx file set as the TFP
# the code checks for a file named after the temp_TFP and deletes it
# then renames the TFP to the temp_TFP                              ^
# the code copies into the temp_TFP                                 |
# this file gets detected and deleted the next time the code runs   |

# I need to test the code just using the TFP instead of the temp_TFP
# that way a new xlsx file doesnt need to be created each time
# instead of any of the code here, id just have TEMP_TARGET_FILE_PATH = TARGET_FILE_PATH

if SIMPLE_WAY:
    TEMP_TARGET_FILE_PATH = TARGET_FILE_PATH
else:
    # Clear Temp_TFP from previous run
    if os.path.exists(TEMP_TARGET_FILE_PATH) and CLEAR_OLD:
        os.remove(TEMP_TARGET_FILE_PATH)

    #Check if we're skipping the copy step
    if SKIP_COPY:
        os.rename(TARGET_FILE_PATH,TEMP_TARGET_FILE_PATH)
        #TEMP_TARGET_FILE_PATH = TARGET_FILE_PATH
    else:
        # Save Target File as a temp file for modification
        copy_xlsx_file(TARGET_FILE_PATH, TEMP_TARGET_FILE_PATH)

# iterate through the sheets of the file
result = iterate_through_sheets(TEMP_TARGET_FILE_PATH)
print("\n All Sheets Read")
if result is not None:
    print("output was valid")
    list_excluded_pages, list_page_names = result

    # Convert the .xlsx file to a pdf
    print(f"Attempting to convert '{TEMP_TARGET_FILE_PATH}' into a pdf file with excel...")

    # Provide the full path to the soffice executable
    raw_pdf_path = xlsx_to_pdf_with_excel(TEMP_TARGET_FILE_PATH, "Masterfile.pdf")

    print(f" Created file: '{raw_pdf_path}'")
    
    #Add pictures to the raw pdf
    print(f"Adding images...")
    
    image_pdf_path = "Logo_"+os.path.basename(raw_pdf_path)
    paste_image_into_pdf(raw_pdf_path, BNR_LOGO_IMAGE_PATH, 40, 0, 145, 145, image_pdf_path)
    print(f" Added Logo: \t\t\t{image_pdf_path}")
    
    image_pdf_path2 = "Signed_"+os.path.basename(raw_pdf_path)
    paste_image_into_pdf(image_pdf_path, SIGNATURE_IMAGE_PATH, 250, 540, 500, 700, image_pdf_path2)
    print(f" Added Signature: \t\t{image_pdf_path2}")
    
    # Set up name for path of final
    FINAL_OUTPUT_PATH = os.path.basename(raw_pdf_path) # optional
    YearAndstrMonth = SOURCE_SHEET_NAME[-4:]+" "+SOURCE_SHEET_NAME[:-5]
    YearAndintMonth = SOURCE_SHEET_NAME[-4:]+" M"+str(MonthStr2Int(SOURCE_SHEET_NAME[:-5])).zfill(2)
    # print(f"YearAndstrMonth: {YearAndstrMonth}") # debug
    # print(f"YearAndintMonth: {YearAndintMonth}") # debug
    YearAndMonth = YearAndintMonth
    
    os.remove(raw_pdf_path)
    print(f" Deleted: {raw_pdf_path}")
    
    dated_name = os.path.splitext(FINAL_OUTPUT_PATH)[0]+" "+YearAndMonth+" Week "+str(WEEK_NUMBER)+" "+current_datetime
    
    final_reports_path = "Final Reports "+YearAndMonth
    os.makedirs(final_reports_path, exist_ok=True)
    FINAL_OUTPUT_PATH = final_reports_path+"/"+dated_name+".pdf"
    
    #Then convert the pdf file into one that doesnt have the excluded pages
    print(f"Creating FINAL_OUTPUT_PATH...")
    pdf_to_pdf_exclude_pages(image_pdf_path2, FINAL_OUTPUT_PATH, list_excluded_pages)

    print(f" Created FINAL_OUTPUT_PATH: \t{FINAL_OUTPUT_PATH}")
    
    
    # conditionally compress pdf
    if PDF_COMPRESS:
        print("COMPRESS")
        compress_pdf(FINAL_OUTPUT_PATH, FINAL_OUTPUT_PATH)
                
    print(f"\nSplitting FINAL_OUTPUT_PATH into Individual Reports...")
    # Then split each page of that pdf into their own pdfs and label them
    # split_pdf_pages(image_pdf_path, list_page_names) # not needed
    
    # okay i have a list of pages names,
    # and a list of integers that are indexes to be deleted from the first list
    
    # print(f"list_page_names:\n{list_page_names}") # debug
    # print(f"list_excluded_pages:\n{list_excluded_pages}") # debug
    
    # this loop deletes from high to low to avoid disturbing the list as we go
    exclude_counter = len(list_excluded_pages)
    while exclude_counter > 0:
        exclude_counter = exclude_counter-1
        # print(f"exclude_counter: {exclude_counter}")
        # print(f"list_excluded_pages[exclude_counter]: {list_excluded_pages[exclude_counter]}")
        # print(f"DELETE: list_page_names[(list_excluded_pages[exclude_counter]-1)]: {list_page_names[(list_excluded_pages[exclude_counter]-1)]}")
        # print(f"DELETE: list_page_names[(list_excluded_pages[exclude_counter])]: {list_page_names[(list_excluded_pages[exclude_counter])]}")
        del list_page_names[(list_excluded_pages[exclude_counter]-1)] # gotta have that -1
    
    # print(f"list_page_names:\n{list_page_names}") # debug
    
    #construct folder name prefix for split_pdf_pages()
    individual_prefix = "Individual Reports "+YearAndMonth+"/"+dated_name
    
    # Then split each page of that pdf into their own pdfs and label them
    split_pdf_pages(individual_prefix, FINAL_OUTPUT_PATH, list_page_names)
    
    # save folder of small pdfs to folder labeled by month and year
    print(f" Saved to folder: \t"+individual_prefix) #debug
    
    if (DELETE_EXTRA):
        # Delete extra files
        print(f"\nDeleting extra files...")
        os.remove(image_pdf_path)
        print(f" Deleted: {image_pdf_path}")
        os.remove(image_pdf_path2)
        print(f" Deleted: {image_pdf_path2}")
        
else:
    print(style.RED + "OUTPUT INVALID" + style.RESET)

###   --Outro

print(f"\n{MY_NAME} END\n")
if len(WarningsList) > 0:
    print("Warnings:"+ style.YELLOW)
    for warning in WarningsList:
        print(warning)
    print(style.RESET)
# Prompt the user to press Enter before closing
input("Press Enter to close...")